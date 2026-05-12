import tkinter as tk
import tkinter.font as tkfont
from tkinter import messagebox, ttk
import openpyxl
from datetime import datetime
import os

# ============================================================
#  การตั้งค่า
# ============================================================
PRODUCTS_FILE = "products.xlsx"
SALES_FILE    = "sales.xlsx"

BG_DARK       = "#0f172a"
BG_CARD       = "#1e293b"
BG_INPUT_MODE = "#0f2027"
BG_SELL_MODE  = "#0a1628"
ACCENT_GREEN  = "#22c55e"
ACCENT_BLUE   = "#3b82f6"
ACCENT_RED    = "#ef4444"
ACCENT_YELLOW = "#f59e0b"
TEXT_WHITE    = "#f8fafc"
TEXT_GRAY     = "#94a3b8"

PREFERRED_FONTS = ["Tahoma", "Segoe UI", "Angsana New", "TH Sarabun New"]
ITEM_TYPES = ["อาหาร", "เครื่องดื่ม", "ขนม", "ของใช้", "เครื่องเขียน", "อื่น ๆ"]


# ============================================================
#  ไฟล์ Excel
# ============================================================
def load_products():
    products = {}
    if not os.path.exists(PRODUCTS_FILE):
        # สร้างไฟล์เปล่าพร้อม header
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "สินค้า"
        ws.append(["บาร์โค้ด", "ชื่อสินค้า", "ราคา", "ประเภท"])
        wb.save(PRODUCTS_FILE)
        return products
    wb = openpyxl.load_workbook(PRODUCTS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            products[str(row[0]).strip()] = {
                "name": str(row[1]) if row[1] is not None else "",
                "price": float(row[2]) if row[2] is not None else 0.0,
                "type": str(row[3]) if len(row) > 3 and row[3] is not None else ""
            }
    return products


def save_product(barcode, name, price, item_type=""):
    if os.path.exists(PRODUCTS_FILE):
        wb = openpyxl.load_workbook(PRODUCTS_FILE)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        if len(header) < 4:
            ws.cell(row=1, column=4, value="ประเภท")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "สินค้า"
        ws.append(["บาร์โค้ด", "ชื่อสินค้า", "ราคา", "ประเภท"])
    # ตรวจว่ามีบาร์โค้ดนี้แล้วหรือยัง → อัปเดต
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).strip() == str(barcode).strip():
            row[1].value = name
            row[2].value = price
            if len(row) < 4:
                ws.cell(row=row[0].row, column=4, value=item_type)
            else:
                row[3].value = item_type
            wb.save(PRODUCTS_FILE)
            return "updated"
    ws.append([barcode, name, price, item_type])
    wb.save(PRODUCTS_FILE)
    return "added"


def save_sale(cart, total):
    if os.path.exists(SALES_FILE):
        wb = openpyxl.load_workbook(SALES_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ยอดขาย"
        ws.append(["วันที่/เวลา", "บาร์โค้ด", "ชื่อสินค้า", "จำนวน", "ราคา/หน่วย", "รวม"])

    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    for barcode, item in cart.items():
        ws.append([
            now,
            barcode,
            item["name"],
            item["qty"],
            item["price"],
            item["price"] * item["qty"]
        ])
    ws.append(["", "", "", "", "ยอดรวม", total])
    ws.append([])
    wb.save(SALES_FILE)


# ============================================================
#  แอปหลัก
# ============================================================
class POSApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ระบบขายสินค้า")
        self.root.configure(bg=BG_DARK)
        self.root.state("zoomed")
        self.root.minsize(900, 600)

        self.window_mode = "zoomed"
        self.font_scale = 1.0
        self.fonts = {}
        self.font_bases = {}
        self.lbl_font_size = None
        self.font_family = self._get_font_family()
        self._init_fonts()
        self.item_types = ITEM_TYPES.copy()

        self.products = load_products()
        self.cart = {}
        self.current_mode = None  # "input" or "sell"

        self._build_home()

    def _init_fonts(self):
        self.fonts = {}
        self.font_bases = {}

    def _get_font_family(self):
        available = set(tkfont.families())
        for name in PREFERRED_FONTS:
            if name in available:
                return name
        return PREFERRED_FONTS[0]

    def _font(self, size, weight="normal"):
        key = (size, weight)
        if key not in self.fonts:
            self.font_bases[key] = size
            self.fonts[key] = tkfont.Font(family=self.font_family,
                                          size=int(size * self.font_scale),
                                          weight=weight)
        return self.fonts[key]

    def _scale(self, value):
        return int(value * self.font_scale)

    def _change_font_scale(self, delta):
        new_scale = round(self.font_scale + delta, 2)
        new_scale = max(0.7, min(1.6, new_scale))
        if new_scale == self.font_scale:
            return
        self.font_scale = new_scale
        for (base_size, weight), font_obj in self.fonts.items():
            font_obj.configure(size=int(self.font_bases[(base_size, weight)] * self.font_scale))
        if self.lbl_font_size:
            self.lbl_font_size.config(text=f"{int(self.font_scale * 100)}%")
        if self.current_mode == "list":
            self._build_list_mode()
        elif self.current_mode == "sell":
            self._build_sell_mode()
        else:
            self._build_home()

    def _apply_window_mode(self):
        if self.window_mode == "zoomed":
            try:
                self.root.state("zoomed")
            except Exception:
                pass
        else:
            self.root.state("normal")
            self.root.geometry("1200x800")

    def _toggle_window_mode(self):
        self.window_mode = "windowed" if self.window_mode == "zoomed" else "zoomed"
        self._apply_window_mode()
        if self.current_mode == "list":
            self._build_list_mode()
        elif self.current_mode == "sell":
            self._build_sell_mode()
        else:
            self._build_home()

    def _build_font_controls(self, parent, bg_color):
        controls = tk.Frame(parent, bg=bg_color)
        tk.Button(controls, text="A-", font=self._font(14, "bold"),
                  bg=BG_CARD, fg=TEXT_WHITE, relief="flat", cursor="hand2",
                  command=lambda: self._change_font_scale(-0.1),
                  activebackground="#475569", activeforeground=TEXT_WHITE).pack(side="left", padx=(0, 8))
        self.lbl_font_size = tk.Label(controls, text=f"{int(self.font_scale * 100)}%",
                                      font=self._font(14), bg=bg_color, fg=TEXT_GRAY)
        self.lbl_font_size.pack(side="left")
        tk.Button(controls, text="A+", font=self._font(14, "bold"),
                  bg=BG_CARD, fg=TEXT_WHITE, relief="flat", cursor="hand2",
                  command=lambda: self._change_font_scale(0.1),
                  activebackground="#475569", activeforeground=TEXT_WHITE).pack(side="left", padx=(8, 0))

        tk.Button(controls, text=("🪟 Window" if self.window_mode == "zoomed" else "🖥️ Full"),
                  font=self._font(14), bg=BG_CARD, fg=TEXT_WHITE, relief="flat",
                  cursor="hand2", command=self._toggle_window_mode,
                  activebackground="#475569", activeforeground=TEXT_WHITE).pack(side="left", padx=(12, 0))
        return controls

    # ============================================================
    #  หน้าหลัก — เลือกโหมด
    # ============================================================
    def _build_home(self):

        self._clear()
        self.current_mode = None

        wrapper = tk.Frame(self.root, bg=BG_DARK)
        wrapper.place(relx=0.5, rely=0.5, anchor="center")
        controls = self._build_font_controls(self.root, BG_DARK)
        controls.place(relx=1.0, y=20, x=-20, anchor="ne")

        tk.Label(wrapper, text="ระบบขายสินค้า", font=self._font(42, "bold"),
                 bg=BG_DARK, fg=TEXT_WHITE).pack(pady=(0, self._scale(8)))
        tk.Label(wrapper, text="เลือกโหมดการใช้งาน", font=self._font(20),
                 bg=BG_DARK, fg=TEXT_GRAY).pack(pady=(0, self._scale(48)))

        btn_frame = tk.Frame(wrapper, bg=BG_DARK)
        btn_frame.pack()

        # โหมดจัดการสินค้า
        self._mode_card(btn_frame,
            emoji="📦",
            title="จัดการสินค้า",
            subtitle="เพิ่ม แก้ไข ลบสินค้า",
            color=ACCENT_BLUE,
            command=self._build_list_mode
        ).grid(row=0, column=0, padx=24)

        # โหมดขาย
        self._mode_card(btn_frame,
            emoji="🛒",
            title="โหมดขายสินค้า",
            subtitle="สแกนสินค้าและดูราคารวม",
            color=ACCENT_GREEN,
            command=self._build_sell_mode
        ).grid(row=0, column=1, padx=24)

        tk.Label(wrapper, text=f"สินค้าในระบบ: {len(self.products)} รายการ",
                 font=self._font(14), bg=BG_DARK, fg=TEXT_GRAY).pack(pady=(self._scale(40), 0))

    def _mode_card(self, parent, emoji, title, subtitle, color, command):
        card = tk.Frame(parent, bg=BG_CARD, cursor="hand2",
                        relief="flat", bd=0)
        card.configure(width=self._scale(300), height=self._scale(220))
        card.pack_propagate(False)

        tk.Label(card, text=emoji, font=self._font(48),
                 bg=BG_CARD).pack(pady=(self._scale(28), 4))
        tk.Label(card, text=title, font=self._font(20, "bold"),
                 bg=BG_CARD, fg=color).pack()
        tk.Label(card, text=subtitle, font=self._font(13),
                 bg=BG_CARD, fg=TEXT_GRAY, wraplength=self._scale(240)).pack(pady=(4, 0))

        # hover effect
        def on_enter(e):
            card.configure(bg=color)
            for w in card.winfo_children():
                w.configure(bg=color)
        def on_leave(e):
            card.configure(bg=BG_CARD)
            for w in card.winfo_children():
                w.configure(bg=BG_CARD)

        card.bind("<Enter>", on_enter)
        card.bind("<Leave>", on_leave)
        card.bind("<Button-1>", lambda e: command())
        for child in card.winfo_children():
            child.bind("<Enter>", on_enter)
            child.bind("<Leave>", on_leave)
            child.bind("<Button-1>", lambda e: command())

        return card

    # ==========================================================
    #  โหมดจัดการสินค้า
    # ==========================================================
    def _build_list_mode(self):
        self._clear()
        self.current_mode = "list"
        self.list_action = None  # "add", "edit", "delete"
        self.list_data = {}
        self.list_search_var = tk.StringVar()
        self.list_sort_column = "ชื่อสินค้า"
        self.list_sort_reverse = False

        root_frame = tk.Frame(self.root, bg=BG_INPUT_MODE)
        root_frame.pack(fill="both", expand=True)
        self._list_frame = root_frame

        # --- header ---
        header = tk.Frame(root_frame, bg="#0c3547", pady=self._scale(12))
        header.pack(fill="x")
        tk.Button(header, text="← กลับ", font=self._font(14),
                  bg="#0c3547", fg=TEXT_GRAY, bd=0, cursor="hand2",
                  activebackground="#0c3547", activeforeground=TEXT_WHITE,
                  command=self._build_home).pack(side="left", padx=20)
        tk.Label(header, text="📦  โหมดจัดการสินค้า",
                 font=self._font(22, "bold"), bg="#0c3547", fg=ACCENT_BLUE).pack(side="left")
        self._build_font_controls(header, "#0c3547").pack(side="right", padx=20)

        # --- body ---
        body = tk.Frame(root_frame, bg=BG_INPUT_MODE)
        body.pack(fill="both", expand=True, padx=self._scale(24), pady=self._scale(16))

        # Left: Product list
        left = tk.Frame(body, bg=BG_CARD, padx=self._scale(24), pady=self._scale(24))
        left.pack(side="left", fill="both", expand=True)

        tk.Label(left, text="รายการสินค้า", font=self._font(20, "bold"),
                 bg=BG_CARD, fg=TEXT_WHITE).pack(pady=(0, self._scale(16)))

        search_frame = tk.Frame(left, bg=BG_CARD)
        search_frame.pack(fill="x", pady=(0, self._scale(12)))
        tk.Label(search_frame, text="ค้นหา:", font=self._font(16),
                 bg=BG_CARD, fg=TEXT_WHITE).pack(side="left")
        self.entry_search = tk.Entry(search_frame, textvariable=self.list_search_var,
                                     font=self._font(16), bg=BG_DARK, fg=TEXT_WHITE,
                                     insertbackground=TEXT_WHITE, relief="flat")
        self.entry_search.pack(side="left", fill="x", expand=True, padx=(8, 0))
        self.entry_search.bind("<KeyRelease>", lambda event: self._refresh_product_list())
        self.lbl_sort_status = tk.Label(search_frame,
                                        text=f"เรียงตาม: {self.list_sort_column}",
                                        font=self._font(14), bg=BG_CARD, fg=TEXT_GRAY)
        self.lbl_sort_status.pack(side="right")

        # Treeview for products
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("List.Treeview",
                        background=BG_DARK,
                        foreground=TEXT_WHITE,
                        fieldbackground=BG_DARK,
                        rowheight=self._scale(40),
                        font=self._font(15))
        style.configure("List.Treeview.Heading",
                        background="#1e3a5f",
                        foreground=TEXT_WHITE,
                        font=self._font(15, "bold"))
        style.map("List.Treeview", background=[("selected", "#1e40af")])

        style.configure("List.TCombobox",
                        fieldbackground=BG_DARK,
                        background=BG_DARK,
                        foreground=TEXT_WHITE,
                        arrowcolor="white",   # 👈 makes arrow visible
                        arrowsize=20,         # 👈 optional: bigger arrow
                        font=self._font(18))
        style.map("List.TCombobox",
                  fieldbackground=[("readonly", BG_DARK), ("!disabled", BG_DARK)],
                  foreground=[("readonly", TEXT_WHITE), ("!disabled", TEXT_WHITE)])
        self.root.option_add("*TCombobox*Listbox.font", self._font(18))
        self.root.option_add("*TCombobox*Listbox.Font", self._font(18))

        cols = ("บาร์โค้ด", "ชื่อสินค้า", "ประเภท", "ราคา")
        self.tree_list = ttk.Treeview(left, columns=cols,
                                      show="headings", style="List.Treeview")
        for col in cols:
            self.tree_list.heading(col, text=col,
                                   command=lambda c=col: self._sort_product_list(c))
        self.tree_list.column("บาร์โค้ด", width=self._scale(140), anchor="center")
        self.tree_list.column("ชื่อสินค้า", width=self._scale(220), anchor="w")
        self.tree_list.column("ประเภท", width=self._scale(180), anchor="center")
        self.tree_list.column("ราคา", width=self._scale(100), anchor="center")

        scrollbar = ttk.Scrollbar(left, orient="vertical", command=self.tree_list.yview)
        self.tree_list.configure(yscrollcommand=scrollbar.set)

        self.tree_list.pack(side="left", fill="both", expand=True, padx=(0, self._scale(8)), pady=(0, self._scale(8)))
        scrollbar.pack(side="right", fill="y", pady=(0, self._scale(8)))

        # Action buttons
        btn_frame = tk.Frame(left, bg=BG_CARD)
        btn_frame.pack(fill="x", pady=(0, self._scale(8)))

        tk.Button(btn_frame, text="➕ เพิ่มสินค้าใหม่",
                  font=self._font(14, "bold"),
                  bg=ACCENT_GREEN, fg="white",
                  relief="flat", cursor="hand2", pady=self._scale(6),
                  command=self._list_show_add).pack(fill="x", pady=(0, self._scale(4)))

        tk.Button(btn_frame, text="✏️ แก้ไขสินค้า",
                  font=self._font(14),
                  bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", pady=self._scale(6),
                  command=self._list_show_edit).pack(fill="x", pady=(0, self._scale(4)))

        tk.Button(btn_frame, text="🗑️ ลบสินค้า",
                  font=self._font(14),
                  bg=ACCENT_RED, fg="white",
                  relief="flat", cursor="hand2", pady=self._scale(6),
                  command=self._list_delete).pack(fill="x")

        # Right: Action panel
        right = tk.Frame(body, bg=BG_CARD, padx=self._scale(24), pady=self._scale(24))
        right.pack(side="right", fill="y")

        self.lbl_list_feedback = tk.Label(right, text="เลือกการดำเนินการ",
                                           font=self._font(16),
                                           bg=BG_CARD, fg=TEXT_GRAY,
                                           wraplength=self._scale(250), justify="center")
        self.lbl_list_feedback.pack(pady=(0, self._scale(20)))

        # Add panel
        self.frame_add = tk.Frame(right, bg=BG_CARD)
        tk.Label(self.frame_add, text="เพิ่มสินค้าใหม่", font=self._font(18, "bold"),
                 bg=BG_CARD, fg=TEXT_WHITE).pack(pady=(0, self._scale(16)))

        tk.Label(self.frame_add, text="บาร์โค้ด:", font=self._font(16), bg=BG_CARD, fg=TEXT_WHITE).pack(anchor="w")
        self.entry_add_barcode = tk.Entry(self.frame_add, font=self._font(18),
                                          bg=BG_DARK, fg=TEXT_WHITE, insertbackground=TEXT_WHITE,
                                          relief="flat", width=self._scale(20))
        self.entry_add_barcode.pack(pady=(0, self._scale(8)))

        tk.Label(self.frame_add, text="ชื่อสินค้า:", font=self._font(16), bg=BG_CARD, fg=TEXT_WHITE).pack(anchor="w")
        self.entry_add_name = tk.Entry(self.frame_add, font=self._font(18),
                                       bg=BG_DARK, fg=TEXT_WHITE, insertbackground=TEXT_WHITE,
                                       relief="flat", width=self._scale(20))
        self.entry_add_name.pack(pady=(0, self._scale(8)))

        tk.Label(self.frame_add, text="ประเภทสินค้า:", font=self._font(16), bg=BG_CARD, fg=TEXT_WHITE).pack(anchor="w")
        self.entry_add_type = ttk.Combobox(self.frame_add, style="List.TCombobox",
                                           font=self._font(18),
                                           values=self.item_types,
                                           state="normal",
                                           postcommand=lambda: self.entry_add_type.config(values=self.item_types),
                                           width=self._scale(20))
        self.entry_add_type.pack(pady=(0, self._scale(8)))

        tk.Label(self.frame_add, text="ราคา (บาท):", font=self._font(16), bg=BG_CARD, fg=TEXT_WHITE).pack(anchor="w")
        self.entry_add_price = tk.Entry(self.frame_add, font=self._font(18),
                                        bg=BG_DARK, fg=TEXT_WHITE, insertbackground=TEXT_WHITE,
                                        relief="flat", width=self._scale(20))
        self.entry_add_price.pack(pady=(0, self._scale(8)))

        tk.Button(self.frame_add, text="💾 เพิ่มสินค้า",
                  font=self._font(16, "bold"),
                  bg=ACCENT_GREEN, fg="white",
                  relief="flat", cursor="hand2", pady=self._scale(8),
                  command=self._list_add_save).pack(pady=(self._scale(16), 0))

        # Edit panel
        self.frame_edit = tk.Frame(right, bg=BG_CARD)
        tk.Label(self.frame_edit, text="แก้ไขสินค้า", font=self._font(18, "bold"),
                 bg=BG_CARD, fg=TEXT_WHITE).pack(pady=(0, self._scale(16)))

        tk.Label(self.frame_edit, text="ชื่อสินค้า:", font=self._font(16), bg=BG_CARD, fg=TEXT_WHITE).pack(anchor="w")
        self.entry_edit_name = tk.Entry(self.frame_edit, font=self._font(18),
                                        bg=BG_DARK, fg=TEXT_WHITE, insertbackground=TEXT_WHITE,
                                        relief="flat", width=self._scale(20))
        self.entry_edit_name.pack(pady=(0, self._scale(8)))

        tk.Label(self.frame_edit, text="ประเภทสินค้า:", font=self._font(16), bg=BG_CARD, fg=TEXT_WHITE).pack(anchor="w")
        self.entry_edit_type = ttk.Combobox(self.frame_edit, style="List.TCombobox",
                                           font=self._font(18),
                                           values=self.item_types,
                                           state="normal",
                                           postcommand=lambda: self.entry_edit_type.config(values=self.item_types),
                                           width=self._scale(20))
        self.entry_edit_type.pack(pady=(0, self._scale(8)))

        tk.Label(self.frame_edit, text="ราคา (บาท):", font=self._font(16), bg=BG_CARD, fg=TEXT_WHITE).pack(anchor="w")
        self.entry_edit_price = tk.Entry(self.frame_edit, font=self._font(18),
                                         bg=BG_DARK, fg=TEXT_WHITE, insertbackground=TEXT_WHITE,
                                         relief="flat", width=self._scale(20))
        self.entry_edit_price.pack(pady=(0, self._scale(8)))

        tk.Button(self.frame_edit, text="💾 บันทึกการเปลี่ยนแปลง",
                  font=self._font(16, "bold"),
                  bg=ACCENT_BLUE, fg="white",
                  relief="flat", cursor="hand2", pady=self._scale(8),
                  command=self._list_edit_save).pack(pady=(self._scale(16), 0))

        self.frame_add.pack_forget()
        self.frame_edit.pack_forget()

        self._refresh_product_list()

    def _refresh_product_list(self):
        # Clear tree
        for row in self.tree_list.get_children():
            self.tree_list.delete(row)

        query = self.list_search_var.get().strip().lower()
        filtered = []
        for barcode, item in self.products.items():
            if query:
                if query not in item["name"].lower() and query not in item["type"].lower():
                    continue
            filtered.append((barcode, item))

        def sort_key(entry):
            barcode, item = entry
            if self.list_sort_column == "บาร์โค้ด":
                return barcode
            if self.list_sort_column == "ชื่อสินค้า":
                return item["name"].lower()
            if self.list_sort_column == "ประเภท":
                return item["type"].lower()
            if self.list_sort_column == "ราคา":
                return item["price"]
            return item["name"].lower()

        for barcode, item in sorted(filtered, key=sort_key, reverse=self.list_sort_reverse):
            self.tree_list.insert("", "end", iid=barcode, values=(
                barcode,
                item["name"],
                item["type"],
                f"{item['price']:.0f} บาท"
            ))

    def _sort_product_list(self, column):
        if self.list_sort_column == column:
            self.list_sort_reverse = not self.list_sort_reverse
        else:
            self.list_sort_column = column
            self.list_sort_reverse = False
        direction = "▼" if self.list_sort_reverse else "▲"
        self.lbl_sort_status.config(text=f"เรียงตาม: {column} {direction}")
        self._refresh_product_list()

    def _list_show_add(self):
        self.list_action = "add"
        self.lbl_list_feedback.config(text="กรอกข้อมูลสินค้าใหม่", fg=ACCENT_GREEN)
        self.frame_add.pack(pady=(self._scale(20), 0))
        self.frame_edit.pack_forget()
        self.entry_add_barcode.delete(0, "end")
        self.entry_add_name.delete(0, "end")
        self.entry_add_type.delete(0, "end")
        self.entry_add_price.delete(0, "end")
        self.entry_add_barcode.focus_set()

    def _list_show_edit(self):
        selected = self.tree_list.selection()
        if not selected:
            self.lbl_list_feedback.config(text="⚠️ กรุณาเลือกสินค้า", fg=ACCENT_RED)
            return
        barcode = selected[0]
        if barcode not in self.products:
            return

        self.list_action = "edit"
        item = self.products[barcode]
        self.list_data = {
            "barcode": barcode,
            "original_name": item["name"],
            "original_price": item["price"],
            "original_type": item.get("type", "")
        }
        self.lbl_list_feedback.config(text=f"แก้ไข: {item['name']}", fg=ACCENT_BLUE)
        self.frame_edit.pack(pady=(self._scale(20), 0))
        self.frame_add.pack_forget()
        self.entry_edit_name.delete(0, "end")
        self.entry_edit_name.insert(0, item["name"])
        self.entry_edit_type.delete(0, "end")
        self.entry_edit_type.insert(0, item.get("type", ""))
        self.entry_edit_price.delete(0, "end")
        self.entry_edit_price.insert(0, str(item["price"]))
        self.entry_edit_name.focus_set()

    def _list_delete(self):
        selected = self.tree_list.selection()
        if not selected:
            self.lbl_list_feedback.config(text="⚠️ กรุณาเลือกสินค้า", fg=ACCENT_RED)
            return
        barcode = selected[0]
        if barcode not in self.products:
            return

        item = self.products[barcode]
        if messagebox.askyesno("ยืนยันการลบ", f"ลบสินค้า: {item['name']} ({barcode}) ?"):
            # Remove from Excel
            wb = openpyxl.load_workbook(PRODUCTS_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value).strip() == barcode:
                    ws.delete_rows(row[0].row)
                    break
            wb.save(PRODUCTS_FILE)
            self.products = load_products()
            self._refresh_product_list()
            self.lbl_list_feedback.config(text="✅ ลบสินค้าแล้ว", fg=ACCENT_GREEN)
            self.frame_add.pack_forget()
            self.frame_edit.pack_forget()
        else:
            self.lbl_list_feedback.config(text="ยกเลิกการลบ", fg=TEXT_GRAY)

    def _list_add_save(self):
        barcode = self.entry_add_barcode.get().strip()
        name = self.entry_add_name.get().strip()
        try:
            price = float(self.entry_add_price.get().strip())
        except ValueError:
            self.lbl_list_feedback.config(text="⚠️ กรุณาใส่ราคาเป็นตัวเลข", fg=ACCENT_RED)
            return

        item_type = self.entry_add_type.get().strip()
        if item_type and item_type not in self.item_types:
            self.item_types.append(item_type)
        if not barcode or not name:
            self.lbl_list_feedback.config(text="⚠️ กรุณากรอกข้อมูลให้ครบ", fg=ACCENT_RED)
            return

        if barcode in self.products:
            self.lbl_list_feedback.config(text="⚠️ บาร์โค้ดนี้มีอยู่แล้ว", fg=ACCENT_RED)
            return

        result = save_product(barcode, name, price, item_type)
        self.products = load_products()
        self._refresh_product_list()
        self.lbl_list_feedback.config(text=f"✅ เพิ่มแล้ว: {name}", fg=ACCENT_GREEN)
        self.frame_add.pack_forget()

    def _list_edit_save(self):
        name = self.entry_edit_name.get().strip()
        try:
            price = float(self.entry_edit_price.get().strip())
        except ValueError:
            self.lbl_list_feedback.config(text="⚠️ กรุณาใส่ราคาเป็นตัวเลข", fg=ACCENT_RED)
            return

        item_type = self.entry_edit_type.get().strip()
        if item_type and item_type not in self.item_types:
            self.item_types.append(item_type)
        if not name:
            self.lbl_list_feedback.config(text="⚠️ กรุณาใส่ชื่อสินค้า", fg=ACCENT_RED)
            return

        result = save_product(self.list_data["barcode"], name, price, item_type)
        self.products = load_products()
        self._refresh_product_list()
        self.lbl_list_feedback.config(text=f"✅ บันทึกแล้ว: {name}", fg=ACCENT_GREEN)
        self.frame_edit.pack_forget()

    # ==========================================================
    #  โหมดขาย
    # ==========================================================
    def _build_sell_mode(self):
        self._clear()
        self.current_mode = "sell"
        self.cart = {}

        root_frame = tk.Frame(self.root, bg=BG_SELL_MODE)
        root_frame.pack(fill="both", expand=True)
        self._sell_frame = root_frame

        # --- header ---
        header = tk.Frame(root_frame, bg="#0a2218", pady=10)
        header.pack(fill="x")
        tk.Button(header, text="← กลับ", font=self._font(14),
                  bg="#0a2218", fg=TEXT_GRAY, bd=0, cursor="hand2",
                  activebackground="#0a2218", activeforeground=TEXT_WHITE,
                  command=self._confirm_back).pack(side="left", padx=20)
        tk.Label(header, text="🛒  โหมดขายสินค้า",
                 font=self._font(22, "bold"), bg="#0a2218", fg=ACCENT_GREEN).pack(side="left")
        self._build_font_controls(header, "#0a2218").pack(side="right", padx=20)

        # --- body: left = scan, right = cart ---
        body = tk.Frame(root_frame, bg=BG_SELL_MODE)
        body.pack(fill="both", expand=True, padx=24, pady=16)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        # ---- LEFT: ช่องสแกน ----
        left = tk.Frame(body, bg=BG_CARD, padx=24, pady=24)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 12))

        tk.Label(left, text="สแกนบาร์โค้ด", font=self._font(20, "bold"),
                 bg=BG_CARD, fg=TEXT_WHITE).pack(pady=(0, self._scale(16)))

        self.entry_scan = tk.Entry(left, font=self._font(26),
                                   bg=BG_DARK, fg=TEXT_WHITE,
                                   insertbackground=TEXT_WHITE,
                                   relief="flat", justify="center")
        self.entry_scan.pack(fill="x", ipady=self._scale(10))
        self.entry_scan.bind("<Return>", self._sell_scan)
        self.entry_scan.focus_set()

        self.lbl_scan_result = tk.Label(left, text="",
                                         font=self._font(18),
                                         bg=BG_CARD, fg=ACCENT_GREEN,
                                         wraplength=self._scale(280), justify="center")
        self.lbl_scan_result.pack(pady=(self._scale(20), 0))

        # ยอดรวม
        tk.Label(left, text="ยอดรวม", font=self._font(18),
                 bg=BG_CARD, fg=TEXT_GRAY).pack(pady=(self._scale(32), 4))
        self.lbl_total = tk.Label(left, text="0 บาท",
                                   font=self._font(40, "bold"),
                                   bg=BG_CARD, fg=ACCENT_YELLOW)
        self.lbl_total.pack()

        # ปุ่ม
        btn_frame = tk.Frame(left, bg=BG_CARD)
        btn_frame.pack(fill="x", pady=(self._scale(32), 0))

        tk.Button(btn_frame, text="✅  ชำระเงิน",
                  font=self._font(18, "bold"),
                  bg=ACCENT_GREEN, fg="white",
                  relief="flat", cursor="hand2", pady=10,
                  activebackground="#16a34a",
                  command=self._checkout).pack(fill="x", pady=(0, 8))

        tk.Button(btn_frame, text="🗑️  ล้างรายการ",
                  font=self._font(16),
                  bg="#334155", fg=TEXT_WHITE,
                  relief="flat", cursor="hand2", pady=8,
                  activebackground="#475569",
                  command=self._clear_cart).pack(fill="x")

        # ---- RIGHT: ตะกร้า ----
        right = tk.Frame(body, bg=BG_CARD)
        right.grid(row=0, column=1, sticky="nsew")

        tk.Label(right, text="รายการสินค้า", font=self._font(20, "bold"),
                 bg=BG_CARD, fg=TEXT_WHITE, pady=self._scale(12)).pack(fill="x", padx=16)

        # Treeview
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("POS.Treeview",
                        background=BG_DARK,
                        foreground=TEXT_WHITE,
                        fieldbackground=BG_DARK,
                        rowheight=self._scale(40),
                        font=self._font(15))
        style.configure("POS.Treeview.Heading",
                        background="#1e3a5f",
                        foreground=TEXT_WHITE,
                        font=self._font(15, "bold"))
        style.map("POS.Treeview", background=[("selected", "#1e40af")])

        cols = ("ชื่อสินค้า", "ราคา", "จำนวน", "รวม")
        self.tree = ttk.Treeview(right, columns=cols,
                                  show="headings", style="POS.Treeview")
        for col in cols:
            self.tree.heading(col, text=col)
        self.tree.column("ชื่อสินค้า", width=self._scale(220), anchor="w")
        self.tree.column("ราคา",       width=self._scale(100), anchor="center")
        self.tree.column("จำนวน",      width=self._scale(80),  anchor="center")
        self.tree.column("รวม",        width=self._scale(110), anchor="center")

        scrollbar = ttk.Scrollbar(right, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=(0, 8))
        scrollbar.pack(side="right", fill="y", pady=(0, 8))

        # ปุ่มลบรายการที่เลือก
        tk.Button(right, text="ลบรายการที่เลือก",
                  font=(FONT, 14),
                  bg=ACCENT_RED, fg="white",
                  relief="flat", cursor="hand2", pady=6,
                  activebackground="#b91c1c",
                  command=self._remove_selected).pack(fill="x", padx=8, pady=(0, self._scale(8)))

    def _sell_scan(self, event=None):
        barcode = self.entry_scan.get().strip()
        self.entry_scan.delete(0, "end")
        if not barcode:
            return

        if barcode not in self.products:
            self.lbl_scan_result.config(
                text=f"⚠️ ไม่พบสินค้า\nบาร์โค้ด: {barcode}", fg=ACCENT_RED)
            return

        item = self.products[barcode]
        if barcode in self.cart:
            self.cart[barcode]["qty"] += 1
        else:
            self.cart[barcode] = {
                "name": item["name"],
                "price": item["price"],
                "qty": 1
            }

        self.lbl_scan_result.config(
            text=f"✅ {item['name']}\n{item['price']:.0f} บาท", fg=ACCENT_GREEN)
        self._refresh_cart()

    def _refresh_cart(self):
        # ล้าง tree
        for row in self.tree.get_children():
            self.tree.delete(row)
        total = 0
        for barcode, item in self.cart.items():
            subtotal = item["price"] * item["qty"]
            total += subtotal
            self.tree.insert("", "end", iid=barcode, values=(
                item["name"],
                f"{item['price']:.0f} บาท",
                item["qty"],
                f"{subtotal:.0f} บาท"
            ))
        self.lbl_total.config(text=f"{total:,.0f} บาท")

    def _remove_selected(self):
        selected = self.tree.selection()
        if not selected:
            return
        for iid in selected:
            if iid in self.cart:
                del self.cart[iid]
        self._refresh_cart()
        self.lbl_scan_result.config(text="ลบรายการแล้ว", fg=TEXT_GRAY)

    def _clear_cart(self):
        if not self.cart:
            return
        if messagebox.askyesno("ยืนยัน", "ล้างรายการทั้งหมด?"):
            self.cart = {}
            self._refresh_cart()
            self.lbl_scan_result.config(text="", fg=ACCENT_GREEN)

    def _checkout(self):
        if not self.cart:
            messagebox.showwarning("แจ้งเตือน", "ยังไม่มีสินค้าในรายการ")
            return
        total = sum(i["price"] * i["qty"] for i in self.cart.values())
        confirm = messagebox.askyesno(
            "ยืนยันการชำระเงิน",
            f"ยอดรวม: {total:,.0f} บาท\n\nบันทึกการขายและล้างรายการ?")
        if confirm:
            save_sale(self.cart, total)
            messagebox.showinfo("สำเร็จ", f"บันทึกการขายแล้ว ✅\nยอดรวม: {total:,.0f} บาท")
            self.cart = {}
            self._refresh_cart()
            self.lbl_scan_result.config(text="", fg=ACCENT_GREEN)
            self.lbl_total.config(text="0 บาท")

    def _confirm_back(self):
        if self.cart:
            if messagebox.askyesno("ยืนยัน", "มีรายการที่ยังไม่ได้ชำระเงิน\nต้องการกลับหน้าหลักใช่ไหม?"):
                self._build_home()
        else:
            self._build_home()

    # ==========================================================
    #  Utility
    # ==========================================================
    def _clear(self):
        for widget in self.root.winfo_children():
            widget.destroy()


# ============================================================
#  เริ่มโปรแกรม
# ============================================================
if __name__ == "__main__":
    root = tk.Tk()
    app = POSApp(root)
    root.mainloop()